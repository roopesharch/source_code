import pyodbc
import pathlib
import time
import teradatasql
import jaydebeapi as jdbc_driver
from socket import gethostname
from datetime import datetime
import pandas as pd
import os
from numpy.ma.core import append
from tabulate import tabulate
from simple_colors import *
import openpyxl
import numpy as np
from robot.libraries.BuiltIn import BuiltIn
import configparser
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Color
# from openpyxl.styles.colors import Font, Color
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import os
import json
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))).replace('\\', '/')
sys.path.append(BASE_DIR)

# Get the Documents folder path
documents_path = os.path.join(os.path.expanduser("~"), "Documents/OPM-AUTOVALIDATION").replace('\\', '/')

def read_json(file_path):
    try:
        with open(file_path, 'r') as f:
            json_data = json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError('Given JSON not found.')
    except json.JSONDecodeError:
        raise ValueError('Invalid JSON format in given file path')
    except Exception as e:
        raise Exception(f'Error reading given file path: {e}')
    return json_data

def read_conf(session, key):
    # print(session, key)
    parser = configparser.ConfigParser()
    # parser.read('C:/Users/RS60/PycharmProjects/OPM_validation/opm-autovalidation/opm_rb/config.ini')
    parser.read(BASE_DIR + '/config.ini')
    return parser[session][key]


# a=read_conf('opm_deliver_stockouts', "stockouts_d_L2_list")
# x = [i for i in a.split(",")]
# print(x)

def set_conf(session, key, value):
    parser = configparser.ConfigParser()
    parser.read(BASE_DIR + '/config.ini')
    parser.set(session, key, value)
    with open(BASE_DIR + '/config.ini', 'w') as config_file_instance:
        parser.write(config_file_instance)

def teradata_connect():
    # Connect to Teradata
    conn = teradatasql.connect(
        host = read_conf('Teradata', 'host'),
        user = read_conf('Teradata', 'user'),
        password = read_conf('Teradata', 'password')
    )
    cursor = conn.cursor()
    return cursor

def connect_kelvin_teradata():
    # conn = pyodbc.connect('Driver={Teradata Database ODBC Driver 16.10};'
    #               'DBCNAME=RATDPRD;'
    #               'AUTHENTICATION=LDAP;'
    #               'UID={rs60};'
    #               "PWD={Freezup@9700000};"
    #               'Trusted_Connection=yes;')

    conn = pyodbc.connect(
        "Driver={" + read_conf('Teradata_connection', 'tera_driver') + "};"
                                                                       "DBCNAME=" + read_conf('Teradata_connection',
                                                                                              'dbcname') + ";"
                                                                                                           "AUTHENTICATION=" + read_conf(
            'Teradata_connection', 'authentication') + ";"
                                                       "UID={" + read_conf('Teradata_connection', 'user_id') + "};"
                                                                                                               "PWD={" + read_conf(
            'Teradata_connection', 'password') + "};"
                                                 "Trusted_Connection=" + read_conf('Teradata_connection',
                                                                                   'trusted_connection') + ";")
    return conn
    # return cursor


# connn = connect_kelvin_teradata()
# print(connn)
# # cursor = connect_synapse()
# ytd_query = pd.read_sql("select top 1 * from CDL.kelvin_npi_otl_history", connn)
# output = pd.DataFrame(ytd_query)
# print(output['Count(*)'][0])
# print("yes done")

def connect_denodo():

    denodoserver_name = read_conf('denodo_connection', 'denodoserver_name')
    denodoserver_jdbc_port = read_conf('denodo_connection', 'denodoserver_jdbc_port')
    denodoserver_database = read_conf('denodo_connection', 'denodoserver_database')
    denodoserver_uid = read_conf('denodo_connection', 'denodoserver_uid')
    denodoserver_pwd = read_conf('denodo_connection', 'denodoserver_pwd')
    denododriver_path = read_conf('denodo_connection', 'denododriver_path')
    denododriver_path = BASE_DIR+'/'+str(denododriver_path)

    client_hostname = gethostname()
    useragent = "%s-%s" % (jdbc_driver.__name__, client_hostname)

    conn_uri = "jdbc:vdb://%s:%s/%s?userAgent=%s" % (denodoserver_name,
                                                     denodoserver_jdbc_port, denodoserver_database,
                                                     useragent)

    cnxn = jdbc_driver.connect("com.denodo.vdp.jdbc.Driver", conn_uri,
                            driver_args={"user": denodoserver_uid, "password": denodoserver_pwd},
                            jars=denododriver_path)

    ## Query to be sent to the Denodo VDP Server
    # query = '''select distinct capacountry  from  "EHS CAPA Data"'''
    cur = cnxn.cursor()
    return cur

# connect_denodo()


def get_color(clr):
    if clr == 'BLACK':
        return '\033[30m'
    elif clr == 'RED':
        return '\033[31m'
    elif clr == 'GREEN':
        return '\033[32m'
    elif clr == 'YELLOW':
        return '\033[33m'
    elif clr == 'BLUE':
        return '\033[34m'
    elif clr == 'MAGENTA':
        return '\033[35m'
    elif clr == 'CYAN':
        return '\033[36m'
    elif clr == 'UNDERLINE':
        return '\033[4m'
    elif clr == 'RESETUNDERLINE':
        return "\033[24m"
    elif clr == 'BLINK':
        return "\033[5m"
    elif clr == 'RESETBOLD':
        return "\033[21m"
    elif clr == 'BOLD':
        return "\033[1m"
    elif clr == 'WHITE':
        return "\033[97m"
    else:
        return '\033[0m'  # reset all


def printf(colr, message):
    BuiltIn().log_to_console("")
    BuiltIn().log_to_console(get_color(colr) + message)
    # print(get_color(colr) + message)


def connect_opm_azure_database():
    sql_connector = pyodbc.connect(
        "Driver=" + read_conf('SQL_connection', 'SQL_drivers') + ";"
                                                                 "Server=" + read_conf('SQL_connection',
                                                                                       'SQL_server') + ";"
                                                                                                       "Database=" + read_conf(
            'SQL_connection', 'SQL_database') + ";"
                                                "UID=" + read_conf('SQL_connection', 'SQL_username') + ";"
                                                                                                       "PWD=" + read_conf(
            'SQL_connection', 'SQL_password') + ";"
                                                "Trusted_Connection=" + read_conf('SQL_connection',
                                                                                  'SQL_Trusted_Connection'))
    return sql_connector.cursor()
    # return sql_connector


def connect_QA_synapse_database():
    sql_connector = pyodbc.connect(
        "Driver=" + read_conf('Synapse_QA', 'SQL_drivers') + ";"
                                                             "Server=" + read_conf('Synapse_QA', 'SQL_server') + ";"
                                                                                                                 "Database=" + read_conf(
            'Synapse_QA', 'SQL_database') + ";"
                                            "UID=" + read_conf('Synapse_QA', 'SQL_username') + ";"
                                                                                               "PWD=" + read_conf(
            'Synapse_QA', 'SQL_password') + ";"
                                            "Trusted_Connection=" + read_conf('Synapse_QA', 'SQL_Trusted_Connection'))
    return sql_connector


def get_scores(lst):
    elements = lst.split('\n')
    # Remove items from element list
    elements = [item for item in elements if item not in ['\n', '\t', '']]
    # BuiltIn().log_to_console(elements)
    if "No scorecard data available" not in elements:
        pass
    else:
        print("No data found")
        BuiltIn().log_to_console("")
        return ['No Data', 'No Data']

    regions = []
    ytd = []
    current_month = []
    for i in range(0, len(elements)):
        item = elements[i]
        print(item)
        if elements[i] == '6MT':
            # regions.append(elements[i-4])
            regions.append(elements[i - 1])

        if elements[i] == 'YTD':
            ytd.append(elements[i + 1])
        if elements[i] == 'CMonth':
            current_month.append(elements[i + 1])
    final_data = {}
    for i, j, z in zip(regions, ytd, current_month):
        #         print(i,j,z)
        l = []
        l.append(j)
        l.append(z)
        final_data[i] = l
    # print(final_data)
    # BuiltIn().log_to_console("Final data")
    # BuiltIn().log_to_console(final_data)
    return final_data


# get_score([1,4,6])

def get_non_year_month_for_sequel_query(date):
    date_str = "'000000',"
    for i in range(int(date[4:6]) + 1, 13):
        zero = ""
        if len(str(i)) == 1:
            zero = "0"
        date_str = date_str + "'" + date[0:4] + zero + str(i) + "',"
    return "(" + date_str[:len(date_str) - 1] + ")"


# d=get_non_year_month_for_sequel_query('202408')
# print(d)

def combine_dictionary(dict1, dict2):
    combined_dict = {key: dict1.get(key, []) + dict2.get(key, []) for key in set(dict1) | set(dict2)}
    return combined_dict

def get_quotient(divident, divisor):
    if divisor != 0:
        return (divident / divisor)
        # print((Number_of_SIFp_w_Highest_Possible_Control / Number_of_SIFp_Reported))
    else:
        return 0


def get_percentage(dividend, divisor):
    if divisor != 0:
        return (dividend / divisor) * 100
        # print((Number_of_SIFp_w_Highest_Possible_Control_YTD/Number_of_SIFp_Reported_YTD))
    elif dividend == 0 and divisor == 0:
        return 100
    else:
        return divisor * 100
        # print(Number_of_SIFp_Reported_YTD)

def combine_dictionary(dict1, dict2):
    combined_dict = {key: dict1.get(key, []) + dict2.get(key, []) for key in set(dict1) | set(dict2)}
    return combined_dict


def delete_empty_sheet(wb, sheet_name):
    # Get the sheet
    sheet = wb[sheet_name]

    # Check if the sheet is empty (no values in any cells)
    if not any(sheet.iter_rows(values_only=True)):  # Check if there are any non-empty rows
        del wb[sheet_name]  # Delete the sheet if it's empty


def write_excel_file(file_path, sheet_name, df, mismatch_score_cell_id, missmatch_l2=[[],[]]):
    # Combine base directory with the file path
    file_path = documents_path + file_path
    # Extract the directory path from the full file path
    directory = os.path.dirname(file_path)
    # Create the directory if it doesn't exist
    if not os.path.exists(directory):
        os.makedirs(directory)
    # Check if the file exists and load it, otherwise create a new workbook
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
    # If 'Sheet' exists and is empty, delete it
    if 'Sheet' in wb.sheetnames:
        delete_empty_sheet(wb, 'Sheet')
    # Check if the sheet already exists
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]  # If the sheet exists, just use the existing one
    else:
        # If the sheet doesn't exist, create a new one
        sheet = wb.create_sheet(sheet_name)
        # Append Header
        sheet.append(list(df))
        # Populating the Excel Sheet with Data
        for row in df.iterrows():
            sheet.append(row[1].tolist())

    # Set column widths for readability
    ws = sheet
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
    ws.column_dimensions = dim_holder
    # Adding mismatch data if provided

    if len(missmatch_l2[0]) != 0:
        sheet.append([''])
        sheet.append(['MISSING L2 CATEGORIES'])
        for i in missmatch_l2[0]:
            sheet.append([i])

    if len(missmatch_l2[1]) != 0:
        sheet.append([''])
        sheet.append(['NEWLY ADDED L2 CATEGORIES'])
        for i in missmatch_l2[1]:
            sheet.append([i])
    # Highlight header with yellow background
    row, col = df.shape
    cell_ids = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']
    # Check link for color code:  https://openpyxl.readthedocs.io/en/stable/styles.html
    temp_font = Font(color="FFFFFF")  # White font color
    header_colour_code = 44
    for i in range(1, col + 1):
        temp = PatternFill(patternType='solid', fgColor=Color(indexed=header_colour_code))  # Yellow for header background
        ws[cell_ids[i - 1] + '1'].fill = temp
        ws[cell_ids[i - 1] + '1'].font = temp_font
    # Highlight mismatch rows if any
    if len(missmatch_l2[0]) != 0 or len(missmatch_l2[1]) != 0:
        for i in range(1, col + 1):
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=header_colour_code))  # Yellow background
            ws[cell_ids[i - 1] + str(row + 3)].fill = temp
            ws[cell_ids[i - 1] + str(row + 3)].font = temp_font
    if len(missmatch_l2[0]) != 0 and len(missmatch_l2[1]) != 0:
        for i in range(1, col + 1):
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=header_colour_code))  # Yellow background
            ws[cell_ids[i - 1] + str(row + len(missmatch_l2[0]) + 5)].fill = temp
            ws[cell_ids[i - 1] + str(row + len(missmatch_l2[0]) + 5)].font = temp_font
    # Highlight mismatch score with yellow
    yellow_clr = 5
    for i in range(1, len(mismatch_score_cell_id) + 1):
        col_id = mismatch_score_cell_id[i - 1]
        temp = PatternFill(patternType='solid',
                           fgColor=Color(indexed=yellow_clr))  # Yellow background for mismatch scores
        ws[col_id].fill = temp
    # Save the workbook
    wb.save(file_path)
    # print('Saved file successfully')

    # Save the workbook
    wb.save(file_path)
    print('Saved file successfully')


def connect_flatfile_azureDB(ytd_query, mtd_query):
    # Connect to Azure database
    cursor = connect_opm_azure_database()

    # Define date range
    start_date = 'Jan-' + str(get_current_date('year'))
    end_date = get_current_date('month_in_word-year')

    # Replace placeholders in queries
    ytd_sql_query = ytd_query.replace('startdate', start_date).replace('enddate', end_date)
    mtd_sql_query = mtd_query.replace('startdate', start_date).replace('enddate', end_date)

    # Execute the queries and store results in DataFrames
    ytd_df = pd.read_sql(ytd_sql_query, cursor)
    mtd_df = pd.read_sql(mtd_sql_query, cursor)

    # Combine results into one DataFrame with labels for clarity
    ytd_df['type'] = 'YTD'
    mtd_df['type'] = 'MTD'
    combined_df = pd.concat([ytd_df, mtd_df], ignore_index=True)

    return combined_df


def write_dictionary_to_data_frame(final_data_dict, column_names, print_df=False, df_append='0', score_card_title='0'):
    # logger.info('Preparing stock out data frame')
    # column_names = ["Date", "Regions", "YTD (OPM)", "Monthly (OPM)", "YTD (synapse)", "Monthly (Synapse)"]
    lst = []
    if isinstance(final_data_dict, str): final_data_dict = json.loads(final_data_dict)

    for i, j in final_data_dict.items():
        k = []
        score_extracted_date = get_current_date('year_month_in_word')
        k.append(score_extracted_date)
        if score_card_title != '0':
            k.append(score_card_title)
        k.append(i)

        # swap scores for arranging  ['a','b','c','d'] -> ['a', 'c', 'b', 'd']
        swap_length = int(len(final_data_dict[i]) / 2)
        count = 0
        for itering in range(swap_length - 1):
            item = final_data_dict[i][count + itering + 2]
            del final_data_dict[i][count + itering + 2]
            final_data_dict[i].insert(itering + 1, item)
            count += 1

        for f in final_data_dict[i]:
            k.append(f)
        lst.append(k)
    # lst.append('')

    # if df_append == '0':
    if isinstance(df_append, pd.DataFrame):

        for i in lst:
            # print(df_append)
            # print(i)
            df_append.loc[len(df_append)] = i
        df = df_append
    else:
        df = pd.DataFrame(lst, columns=column_names)
        df.columns = df.columns.str.upper()
    # print(df)

    if score_card_title == '0':
        table = tabulate(df, headers="keys", tablefmt="fancy_grid", showindex="always")
        df.style.highlight_max()
        colored_table = green(table)  # You can use other color functions like red(), blue(), etc.
        BuiltIn().log_to_console("")
        # BuiltIn().log_to_console(color_text.BLUE +"Data table")
        BuiltIn().log_to_console(get_color("UNDERLINE") + "Data table")
        BuiltIn().log_to_console(get_color("RESETUNDERLINE") + "")
        temp_data = read_json(BASE_DIR + '/temp.json')
        KPI_name = temp_data['kpiName']
        BuiltIn().log_to_console(KPI_name)
        BuiltIn().log_to_console(colored_table)

    return df


# final_data_dict = {'APAC': [0, 0, 0, 0], 'WW': [3, 0, 3, 0], 'NA': [0, 0, 0, 0], 'EMEA': [0, 0, 0, 0], 'LATAM': [3, 0, 3, 0]}
#
# d=write_dictionary_to_data_frame(final_data_dict,["Date", "Regions", "YTD (OPM)", "Monthly (OPM)", "YTD (synapse)", "Monthly (Synapse)"],True)
# print(d)


def get_month_name_in_words(date):
    month_data = {'01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr', '05': 'May', '06': 'Jun', '07': 'Jul',
                  '08': 'Aug', '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dec'}
    return month_data[date]


def set_current_date():
    today = str(datetime.today())
    # print(today)
    dates = today.split('-')
    # print(dates)

    if str(dates[1]) == '01' or str(dates[1]) == '1':
        month = '12'
    else:
        month = str((int(dates[1]) - 1))
        if len(month) == 1:
            month = '0' + month
    if str(dates[1]) == '01' or str(dates[1]) == '1':
        year = str(int(dates[0]) - 1)
    else:
        year = str(dates[0])

    return year + month


# set_current_date()

def get_current_date(get_date):
    # today = read_conf('opm', 'year_month')
    temp_data = read_json(BASE_DIR + '/temp.json')
    today = temp_data['filters']['date']
    # now = data[]
    year = today[:4]
    month = today[4:]
    month_in_word = get_month_name_in_words(month)

    if get_date == 'month':
        return month
    elif get_date == 'month_in_word':
        return month_in_word
    elif get_date == 'year':
        return year
    elif get_date == 'yearmonth':
        return str(year) + str(month)
    elif get_date == 'monthyear':
        return str(month) + str(year)
    elif get_date == 'year-month':
        return str(year) + '-' + str(month)
    elif get_date == 'year_month_in_word':
        return str(year) + '_' + str(month_in_word)
    elif get_date == 'month_in_word-year':
        return month_in_word + '-' + str(year)
    return


# da=get_current_date('month_in_word')
# print(da)


def Compare_lists(old_l2_lst, score_dict):
    # old_l2_lst = old_l2_lst.split(",")
    missing_l2 = []
    new_l2 = []
    if isinstance(score_dict, str): score_dict = json.loads(score_dict)

    for j in old_l2_lst:
        if str(j) not in score_dict:
            missing_l2.append(j)

    for i in score_dict:
        if i not in old_l2_lst:
            new_l2.append(i)

    if len(missing_l2) == 0 and len(new_l2) == 0:
        # BuiltIn().log_to_console(get_color('YELLOW')+" Old and new list match correctly")
        return [missing_l2, new_l2, 'No error']
    else:
        if len(missing_l2) != 0:
            # BuiltIn().log_to_console(get_color('RED')+ "Missing L2 from scorecard ")
            missing_l2_df = pd.DataFrame(missing_l2, columns=['Missing L2 List From Scorecard'])
            missing_l2_df.columns = missing_l2_df.columns.str.upper()
            missing_l2_df_table = tabulate(missing_l2_df, headers="keys", tablefmt="fancy_grid", showindex="always")
            missing_l2_df_table = red(missing_l2_df_table)
            BuiltIn().log_to_console('')
            BuiltIn().log_to_console(missing_l2_df_table)

        if len(new_l2) != 0:
            # BuiltIn().log_to_console(get_color('RED')+ "Newly added L2 into scorecard  ")
            # new_l2_df = pd.DataFrame(new_l2)
            new_l2_df = pd.DataFrame(new_l2, columns=['Newly added L2 into scorecard'])
            new_l2_df.columns = new_l2_df.columns.str.upper()
            new_l2_df_table = tabulate(new_l2_df, headers="keys", tablefmt="fancy_grid", showindex="always")
            new_l2_df_table = green(new_l2_df_table)

            # missing_l2_df = pd.DataFrame(missing_l2)
            # BuiltIn().log_to_console('')
            BuiltIn().log_to_console(new_l2_df_table)

    return [missing_l2, new_l2]


# final_data_dict = {'APAC1': [0, 0, 0, 0],'APAC': [0, 0, 0, 0], 'WW': [3, 0, 3, 0], 'NA': [0, 0, 0, 0], 'EMEA': [0, 0, 0, 0], 'LATAM': [3, 0, 3, 0]}
# print(verify_scores( ['APAC', 'WW', 'NA', 'EMEA', 'LATAM','maz'] ,final_data_dict))

def compare_scores(df, column_to_ignore=2, kpi_name=' '):
    cell_id_dict = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L',
                    13: 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W',
                    24: 'X', 25: 'Y', 26: 'Z'}

    mismatch_score_cell_id = []
    unmatched_l2 = []
    header = list(df.columns)
    row, column = df.shape
    # source_count = int((len(header)-4)/2)
    source_count = int((len(header) - (2 + column_to_ignore)) / 2)
    # print("source_count",source_count)
    error = ''
    number_of_compare = 2  # eg: ytd, monthly = 2 count
    # column_to_ignore = date and L2 columns  = 2  as default
    for i in range(column_to_ignore, column_to_ignore + number_of_compare + 2, 2):
        # print("i:",i)
        for j in range(0, row):
            # print("j:",j)
            # compare_count = i

            for k in range(0, source_count):
                # compare_count = compare_count + 2
                opm_score = df[header[i]][j]
                source_score = df[header[i + 1]][j]

                # Score type conversion
                # Convert to float if source is float type else round to integer
                if opm_score in ['TBD', 'NA', 'No Data', 'N/A']:
                    pass
                elif '%' in str(opm_score):
                    opm_score = opm_score.split('%')[0] if '.' in str(opm_score) and '.' in str(source_score) else str(
                        round(float(opm_score.split('%')[0])))
                else:
                    opm_score = opm_score if '.' in str(opm_score) and '.' in str(source_score) else str(
                        round(float(opm_score)))

                if source_score in ['TBD', 'NA', 'No Data', 'N/A']:
                    pass
                elif '%' in str(source_score):
                    source_score = source_score.split('%')[0] if '.' in str(source_score) and '.' in str(
                        opm_score) else str(round(float(source_score.split('%')[0])))
                else:
                    source_score = source_score if '.' in str(source_score) and '.' in str(opm_score) else str(
                        round(float(source_score)))


                temp_data = read_json(BASE_DIR + '/temp.json')
                tolerance = temp_data['tolerance']

                if source_score or opm_score in ['TBD', 'NA', 'No Data', 'N/A']:
                    if opm_score != source_score:
                        mismatch_score_cell_id.append(str(cell_id_dict[i + 1]) + str(j + 2))
                        if header[i] not in unmatched_l2:
                            unmatched_l2.append(header[i] + " : " + df[header[column_to_ignore - 1]][j])
                elif abs(float(opm_score)-float(source_score)) <= float(tolerance):
                    pass
                else:
                    mismatch_score_cell_id.append(str(cell_id_dict[i + 1]) + str(j + 2))
                    if header[i] not in unmatched_l2:
                        unmatched_l2.append(header[i] + " : " + df[header[column_to_ignore - 1]][j])

    print(mismatch_score_cell_id, unmatched_l2)
    if str(error) != '' and "TBD" in str(error):
        BuiltIn().log_to_console('')
        BuiltIn().log_to_console(get_color('RED') + 'Score Not Available, To Be Decided')
        # BuiltIn().log_to_console(error)
    elif str(error) != '' and "NA" in str(error):
        BuiltIn().log_to_console('')
        BuiltIn().log_to_console(get_color('RED') + 'Score Not Available, Not Available')

        BuiltIn().log_to_console(get_color('RED') + "KPI : " + str(kpi_name))
        unmatched_l2_df = pd.DataFrame(unmatched_l2, columns=['L2 with incorrect score match'])
        unmatched_l2_df.columns = unmatched_l2_df.columns.str.upper()
        unmatched_l2_df_table = tabulate(unmatched_l2_df, headers="keys", tablefmt="fancy_grid", showindex="always")
        unmatched_l2_df_table = red(unmatched_l2_df_table)
        BuiltIn().log_to_console('')
        BuiltIn().log_to_console(unmatched_l2_df_table)
    return mismatch_score_cell_id


def write_multiple_json_to_dataframe(final_data_dict, column_names):
    df_dictionary = {}
    if isinstance(final_data_dict, str): final_data_dict = json.loads(final_data_dict)

    for i in final_data_dict:
        ind = i.index(':')
        region_l2 = i[:ind]
        scorecard_data = i[ind + 1:]

        if scorecard_data not in df_dictionary.keys():
            df = pd.DataFrame([], columns=column_names)
            df.columns = df.columns.str.upper()
            df_dictionary[scorecard_data] = df

        # create a dataframe with only header and try to append data in instalements to same DF as its a nested dictionary
        if isinstance(final_data_dict[i], list):
            temp_dic = {'L2 Not Available': final_data_dict[i]}
            df = write_dictionary_to_data_frame(temp_dic, column_names, False, df_dictionary[scorecard_data], region_l2)
        else:
            df = write_dictionary_to_data_frame(final_data_dict[i], column_names, False, df_dictionary[scorecard_data],
                                                region_l2)
        df_dictionary[scorecard_data] = df

    # BuiltIn().log_to_console(df_dictionary)
    # print(df_dictionary)

    for i, j in df_dictionary.items():
        # print(j)
        table = tabulate(j, headers="keys", tablefmt="fancy_grid", showindex="always")
        j.style.highlight_max()
        colored_table = green(table)  # You can use other color functions like red(), blue(), etc.
        # print(i)
        # print(colored_table)

        BuiltIn().log_to_console("")
        # BuiltIn().log_to_console(color_text.BLUE +"Data table")
        BuiltIn().log_to_console(get_color("UNDERLINE") + "Data table")
        BuiltIn().log_to_console(get_color("RESETUNDERLINE") + "")

        # color_text.BLUE + "Hello, World!"
        BuiltIn().log_to_console(i)
        BuiltIn().log_to_console(colored_table)

    print("end ")
    return df_dictionary


def compare_multiple_kpi_score_df(df_dic):
    error_count = 0
    df_and_wrong_score_dic = {}
    for i, j in df_dic.items():
        wrong_score_list = compare_scores(j, 3, i)
        df_and_wrong_score_dic[i] = [j, wrong_score_list]
        if len(wrong_score_list) > 1:
            error_count = 5
    if error_count == 5:
        return [df_and_wrong_score_dic, 5]
    return [df_and_wrong_score_dic]


def write_multiple_KPI_df_to_excel_report(file_path, df, missmatch_l2=[[],[]]):

    for i, j in df.items():
        if len(missmatch_l2[0]) == 0 and len(missmatch_l2[1]) == 0:
            write_excel_file(file_path, str(i)[:30], df[i][0], df[i][1])
        else:
            write_excel_file(file_path, str(i)[:30], df[i][0], df[i][1], missmatch_l2)

# def get_quotient(divident, divisor):
#     if divisor != 0:
#         return (divident / divisor)
#         # print((Number_of_SIFp_w_Highest_Possible_Control / Number_of_SIFp_Reported))
#     else:
#         return 0


# def get_percentage(dividend, divisor):
#     if divisor != 0:
#         return (dividend / divisor) * 100
#         # print((Number_of_SIFp_w_Highest_Possible_Control_YTD/Number_of_SIFp_Reported_YTD))
#     elif dividend == 0 and divisor == 0:
#         return 100
#     else:
#         return divisor * 100
        # print(Number_of_SIFp_Reported_YTD)





# dics = {"WW:Severe Injury or Fatality - Precursor (SIF-P)": ["No Data", "No Data", "No Data", "No Data"], "WW:Severe Injury or Fatality - Precursor (SIF-P) Reporting": {"WW": ["N/A", "N/A", 0, 0], "NA": ["N/A", "N/A", 0, 0], "LATAM": ["N/A", "N/A", 0, 0], "EMEA": ["N/A", "N/A", 0, 0], "APAC": ["N/A", "N/A", 0, 0]}, "WW:Severe Injury or Fatality - Controls (SIF-P Controls)": {"WW": ["N/A", "N/A", 97.12879409351928, 0], "NA": ["N/A", "N/A", 97.48201438848922, 0], "LATAM": ["N/A", "N/A", 0, 0], "EMEA": ["N/A", "N/A", 97.93577981651376, 0], "APAC": ["N/A", "N/A", 0, 0]}, "LATAM:Severe Injury or Fatality - Precursor (SIF-P)": ["No Data", "No Data", "No Data", "No Data"], "LATAM:Severe Injury or Fatality - Precursor (SIF-P) Reporting": {"LATAM": ["N/A", "N/A", 0, 0]}, "LATAM:Severe Injury or Fatality - Controls (SIF-P Controls)": {"LATAM": ["N/A", "N/A", 0, 0]}, "EMEA:Severe Injury or Fatality - Precursor (SIF-P)": ["No Data", "No Data", "No Data", "No Data"], "EMEA:Severe Injury or Fatality - Precursor (SIF-P) Reporting": {"EMEA": ["N/A", "N/A", 0, 0]}, "EMEA:Severe Injury or Fatality - Controls (SIF-P Controls)": {"EMEA": ["N/A", "N/A", 97.93577981651376, 0]}}
#
# column_names = ["Date", "Site", "L2_Regions", "YTD (OPM)", "YTD (Denodo)", "Monthly (OPM)", "Monthly (Denodo)"]
#
# d=write_multiple_json_to_dataframe(dics, column_names)
# print(d)
# print("yes")
# f = compare_multiple_kpi_score_df(d)
#
# file_path="/score_report/ehs_KPI.xlsx"
# missmatch_l2=[[],[]]
# write_multiple_KPI_df_to_excel_report(file_path, f[0], missmatch_l2)


# final_data_frame = write_dictionary_to_data_frame({'LM & SM': ['98.5%', '100.0%', '98%', '100%']}, ['Date', 'L2', 'YTD (OPM)', 'YTD (Source)', 'Monthly (OPM)', 'Monthly (Source)'])
# print(final_data_frame)
# write_excel_file("/score_report/ehs/new_data.xlsx", "score_sheet_new", final_data_frame, ['E2', 'F2'], [[],[]])
